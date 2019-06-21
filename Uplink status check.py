"""
Written for Python 3.7.2
Created by Nate Revello 2019
roughly 30.5 hours as of 2/1 3pm

-We’ll have to use the windows task scheduler to run once a day (if not more often).  If a device has been down for over the threshold, then it'll be added to the report that's emailed out.

# curl setup: paste curl.exe, curl-ca-bundle.crt, and libcurl-x64.dll, into system32
"""

# Once working, create a Meraki Script account.  replace API keys and whatnot.

import os # for checking if the excel file already exists, etc.
import subprocess # for running cmd commands
from time import sleep # for pauses.  API only approves 5 calls per sec.
from datetime import datetime # for timestamps
from datetime import timedelta
from openpyxl import Workbook # a excel file will be used to store some data to compare timestamps with and send an alert if an interface has been down for x amount of time.  An excel file would be easier for checking rows/columns vs a csv [in my opinion].
from openpyxl import load_workbook # used for seeing if the excel file already exists


#################################################################################
#################################################################################
###																			  ###
###					INSERT THE API KEY BELOW								  ###
###						Modify filenames as needed							  ###
###						Edit the email info in the other email script		  ###
###																			  ###
#################################################################################
#################################################################################
####	This is the Meraki API Key.  Use the one for your account, or setup an API/dev account
api_key = "###"
#################################################################################
#################################################################################
####	This is the xlsx file for tracking uplinks and checking their timestamps
uplink_excel_list = "DO.NOT.MODIFY.Meraki Uplink Script List.xlsx" 
#################################################################################
#################################################################################
####	This is for the xlsx file that'll be emailed in the alert
alert_excel_list = "Meraki Uplink Alert List.xlsx"
#################################################################################
#################################################################################




#.+*^*+.+*^*+.+*^*+.+*^*+.+*^*+.+*^*+.+*^*+.+*^*+.+*^*+.+*^*+.+*^*+.+*^*+.+*^*+.#
#################################################################################
################## Methods ######################################################

def save_xlsx(): # a method to save the workbook
	wb.save(uplink_excel_list) # Save the workbook

def get_row_count(): # a method to get the max row count in the excel file
	row_count = ws.max_row
	return row_count

def get_col_count(): # a method to get the max column count in the excel file
	col_count = ws.max_column
	return col_count



#################################################################################
################## preliminary excel work #######################################
if os.path.isfile(uplink_excel_list) == False: # check if the excel file already exists.  If it doesn't create the file and write the headers.
	print()
	print(uplink_excel_list+" does not exist.  Creating it now")
	wb = Workbook() # create workbook
	ws = wb.active # grab the sheet
	ws.cell(column=1, row=1, value="Device name") # write column headers
	ws.cell(column=2, row=1, value="Serial number")
	ws.cell(column=3, row=1, value="Model")
	ws.cell(column=4, row=1, value="Interface")
	ws.cell(column=5, row=1, value="Int Status")
	ws.cell(column=6, row=1, value="Date Time1")
	ws.cell(column=7, row=1, value="Date Time2") # this second column is to compare with the first.  if the difference is met, then the device info is sent in an alert.
	ws.cell(column=8, row=1, value="Alert?")
	print("Created file and wrote headers...")
	save_xlsx()
else:
	print(uplink_excel_list+" does exist.  Loading it now...")
	wb = load_workbook(uplink_excel_list)
	ws = wb.active
	print("Load complete")
	
#################################################################################
################## Org ID #######################################################
# Pull the network ID.  Convert to string.
#
print("\nPulling organization info...\n")
org_curl = str(subprocess.check_output('curl -L -H "X-Cisco-Meraki-API-Key: '+api_key+'" -H "Content-Type: application/json" -X GET "https://api.meraki.com/api/v0/organizations"')) # output: [{"id":######,"name":"Company Name"}]
#print("org_curl: ",org_curl) # test output: b'[{"id":######,"name":"Company Name"}]'
org_split = org_curl.split(',')
#print("org_split: ",org_split) # output: ['b\'[{"id":######', '"name":"Company Name"}]\'']
#print() # print blank line


clean_org_list = list() # create a new list for cleaned up text
line_list = list() # create a new list for each entry	

for entry in org_split:
	#print("Entry: ",entry) # for testing
	entry = entry.replace("{","") # remove {
	entry = entry.replace("}","") # remove }
	entry = entry.replace("]","") # remove ]
	entry = entry.replace("[","") # remove [
	entry = entry.replace("b'","") # remove b'
	entry = entry.replace('"',"") # remove "

	#print("Entry after no {} and whatnot: ",entry) # used for testing
	clean_org_list.append(entry) #append the new entry to the updated list (no curly braces).
	#line_list = entry.split(',') # split each entry into it's own list.
	
	# create column variables.  blanked out for formatting/testing
	organization_id = 'blank'
	
	#print() # print a clean line
	
	for word in clean_org_list: # go through the line since it's a set
			#print("word: ",word) # output for testing
			#print()
			word_list = list() # create a list for each "word" in the line.
			word_list = word.split(":")
			#print("word_list:",word_list)
			
			if 'id' in word_list:
				network_id = word_list[1]
			#print("id: ",network_id) # for testing
			#print()
	#print("clean_org_list: ",clean_org_list)

####################################################################################
################## Network ID ######################################################
# Scrape the IDs and get the appropriate network tags.  Grab the serial numbers
# curl -L -H 'X-Cisco-Meraki-API-Key: ##############################' -H 'Content-Type: application/json' -X GET 'https://api.meraki.com/api/v0/organizations/######/networks'

clean_network_list = list() # create a new list for cleaned up text
sleep(0.8) # sleep for half a sec
print("\nPulling network info...\n")
network_curl = str(subprocess.check_output('curl -L -H "X-Cisco-Meraki-API-Key: '+api_key+'" -H "Content-Type: application/json" -X GET "https://api.meraki.com/api/v0/organizations/'+network_id+'/networks"')) # output: a ton'o'text
# print("network_curl: ",network_curl) # for testing
split_network_output = network_curl.split("},{") # split the text into a list
# print(split_network_output) # test output

# create column variables.  blanked out for formatting/testing
id = 'blank' # this is the network id (site in meraki).
organization_id = 'blank'
name = 'blank'
time_zone = 'blank'
tags = 'blank'
type = 'blank'
config_template_id = 'blank'
disable_my_meraki_com = 'blank'
disable_remote_status_page = 'blank'

for line in split_network_output: # go through the network list, one entry at a time. line = counter variable

	#print("#################################")
	# print("Line :",line) # test ouput
	line_list = list() # create a new list for each line	
	line_list = line.split(',') # split each line into it's own list.
	#print("line_list: ",line_list) # test output
	#print()
	
	for entry in line_list: # go through the line since it's a set
		#print("line_list loop - Beginning")
		word_list = list() # create a list for each "word" in the line.
		# clean up the un-needed stuff
		#print("entry (pre-cleanup): ",entry)
		
		entry = entry.replace("{","") # remove {
		entry = entry.replace("}","") # remove }
		entry = entry.replace("]","") # remove ]
		entry = entry.replace("[","") # remove [
		entry = entry.replace("b'","") # remove b'
		entry = entry.replace('"',"") # remove "
		entry = entry.replace("'","") # remove '
		word_list = entry.split(':')
		
		#print("entry (post-cleanup: ",entry)
		#print("Word list: ",word_list)
		#print("line_list loop - end")
		#print()
		#print()
		
		# go through and set the variables.  Not all are used, but they're set in case they're needed in the future. 1/22/19 NR
		for word in word_list:
			#print("word: ",word)
			
			if 'id' in word_list:
				id = word_list[1] # set the variable.
				#print("id value: ", id) # this prints the second item in the list--which happens to be the value.  example output: id value:  N_123456789123456789
				#print() #blank line for testing
			if 'organizationId' in word_list:
				organization_id = word_list[1] # set the variable.
				#print("organizationId value: ", organization_id)
			if 'name' in word_list:
				name = word_list[1] # set the variable.
				#print("name value: ", name)
			if 'timeZone' in word_list:
				time_zone = word_list[1] # set the variable.
				#print("timeZone value: ", time_zone)
			if 'tags' in word_list:
				tags = word_list[1].strip() # set the variable and remove whitespace at the beginning and end.
				#print("tags value (Ignore single quotes): '"+tags+"'")
			if 'type' in word_list:
				type = word_list[1] # set the variable.
				#print("type value: ", type)	
			if 'configTemplateId' in word_list:
				config_template_id = word_list[1] # set the variable.
				#print("configTemplateId value: ", config_template_id)		
			if 'disableMyMerakiCom' in word_list:
				disable_my_meraki_com = word_list[1] # set the variable.
				#print("disableMyMerakiCom value: ", disable_my_meraki_com)		
			if 'disableRemoteStatusPage' in word_list:
				disable_remote_status_page = word_list[1] # set the variable.
				#print("disableRemoteStatusPage value: ", disable_remote_status_page)
		#print()

	
	
	
	####################################################################################
	################## SN info #########################################################
	# 	Grab the SN for every 'id'
	# 	Curl: curl -L -H "X-Cisco-Meraki-API-Key: <key>" -H "Content-Type: application/json" -X GET "https://api.meraki.com/api/v0/networks/[networkId]/devices"
	
	#print("##################################################################")
	if tags == "field":
	
		sleep(0.6) # sleep for half a sec
		print("\nPulling serial number & device info..."+name+"\n")
		sn_curl = str(subprocess.check_output('curl -L -H "X-Cisco-Meraki-API-Key: '+api_key+'" -H "Content-Type: application/json" -X GET "https://api.meraki.com/api/v0/networks/'+id+'/devices"'))
		#print()
		#print()
		#print("################  SN Curl  #################")
		
		#print("raw sn_curl: \n",sn_curl) # output example: sn_curl:  b'[{"lanIp":"10.10.100.100","serial":"SN12-3456-7890","mac":"##:##:##:##:##:##","lat":12.34567,"lng":-12.34567,"address":"123 Main St, NY, NY 12345","tags":" Tage are here ","name":"device name","model":"MR42","networkId":"L_123456789123456789"},{"wan1Ip":"192.168.1.4","wan2Ip":"192.168.0.5","lanIp":"192.168.1.5","serial":"SN34-5678-9012","mac":"##:##:##:##:##:##","lat":12.12345,"lng":-12.12345,"address":"1234 main st, ny, ny 12345","name":"mx65 device name","model":"MX65","networkId":"L_123456789123456789"}]'
		#print()
		#print()
		#print()
		
		#########################################################################
		#		start cleaning up text from the device:							#
		#########################################################################
		sn_split = sn_curl.split('},{') # split the output into different sections.  Each section is a device (wap, mx65, etc)
		
		clean_sn_list = list() # create a new list for cleaned up text
		sn_section_number = 0 # create a section variable.  This helps with keeping the devices separate.
		
		for sn_curl_entry in sn_split:
			#print("sn_section_number: ",sn_section_number)
			sn_section_number += 1 # update the section counter
			#print()
			#print()
		
			sn_device_list = list() # create a new list for each entry
			
			#print("sn_curl_entry (pre-cleanup and split):\n",sn_curl_entry) # for testing
			#print() # blank line for testing
			sn_curl_entry = sn_curl_entry.replace("{","") # remove {
			sn_curl_entry = sn_curl_entry.replace("}","") # remove }
			sn_curl_entry = sn_curl_entry.replace("]","") # remove ]
			sn_curl_entry = sn_curl_entry.replace("[","") # remove [
			sn_curl_entry = sn_curl_entry.replace("b'","") # remove b'
			sn_curl_entry = sn_curl_entry.replace('"',"") # remove "
			sn_curl_entry = sn_curl_entry.replace("'","") # remove '
			#print("sn_curl_entry after no {} and whatnot: ",sn_curl_entry) # used for testing
			#print() # blank line for testing
			
			sn_device_list = sn_curl_entry.split(',')
			#print("sn_device_list after cleanup and split:\n",sn_device_list)
			#print("\n\n\n\n")
			
			needed_device_sn = "no" # set the variable for the device SN.  if this is set to "yes" then the SN will be passed to the uplink check.
			
			# clear out variables:
			sn_serial = "blank"
			sn_name = "blank"
			sn_model = "blank"
			
			for sn_device_entry in sn_device_list: # go through the entries and split them
				#print('First four Chars: ',sn_device_entry[0:4])
				if sn_device_entry[0:4] == "mac:": 
					#split the first colon
					#print('MAC Entry.  First four Chars: ',sn_device_entry[0:4])
					sn_device_entry = sn_device_entry.replace('mac:','mac.')
					sn_device_attribute = sn_device_entry.split('.')
				else:
					sn_device_attribute = sn_device_entry.split(':')
				#print("sn_device_attribute: ",sn_device_attribute)
				
				#if len(sn_device_attribute) == 1: # if there's just one item in the list, don't spit out an error.
					#print("sn_device_attribute only has one item: ",sn_device_attribute[0])
				if len(sn_device_attribute) == 2: # if there's two items (ex: name: <value>) in the list, then do the following checks.
					#print("sn_device_attribute[1]: ",sn_device_attribute[1])
					if sn_device_attribute[0] == 'serial': # Grab the SN
						sn_serial = sn_device_attribute[1]
						#print("sn_serial: ",sn_serial)
					if sn_device_attribute[0] == 'name': # Grab the name
						sn_name = sn_device_attribute[1]
						#print("sn_name: ",sn_name)
					if sn_device_attribute[0] == 'model': # begin model check
						sn_model = sn_device_attribute[1]
						#print("sn_model: ",sn_model)
						if sn_model[0:2] == 'MX': # check for MX devices
							#print("needed device") # check for testing
							needed_device_sn = 'yes' # set a variable for the future variable pass (sn for uplink check)
							#print("needed_device_sn (Yes or no): ",needed_device_sn)
						#else: # if not an MX appliance
							#print(sn_model[0:2]+" does not match 'MX'")
							#print("needed_device_sn (Yes or no): ",needed_device_sn)
			#print()
			#print()
		#print() # blank line for testing
		#print() # blank line for testing
	#break # break out of the loop for testing

			####################################################################################
			################## Uplink info #####################################################
			# Insert the network ID, and serial number, into the uplink pull:
			# Uplink request: https://dashboard.meraki.com/api_docs#return-the-uplink-information-for-a-device
			# Command: curl -L -H 'X-Cisco-Meraki-API-Key: <key>' -H 'Content-Type: application/json' -X GET 'https://api.meraki.com/api/v0/networks/[networkId]/devices/[serial]/uplink'
			if needed_device_sn == 'yes':
				#print("Starting Uplink Curl\n")
				sleep(0.6) # wait / don't hit the API too much
				#print("sn_model: ",sn_model)
				#print("sn_serial: ",sn_serial)
				#print("[network] id: ",id)
				#print("api_key: ",api_key)
				#print('command check: '+'curl -L -H "X-Cisco-Meraki-API-Key: '+api_key+'" -H "Content-Type: application/json" -X GET "https://api.meraki.com/api/v0/networks/'+id+'/devices/'+sn_serial+'/uplink"')
				#print("\nPulling uplink info...\n")
				uplink_curl = str(subprocess.check_output('curl -L -H "X-Cisco-Meraki-API-Key: '+api_key+'" -H "Content-Type: application/json" -X GET "https://api.meraki.com/api/v0/networks/'+id+'/devices/'+sn_serial+'/uplink"'))
				#print("\nuplink_curl: \n",uplink_curl) # test output: uplink_curl:  b'[{"interface":"WAN 1","status":"Active","ip":"192.168.1.12","gateway":"192.168.1.254","publicIp":"12.34.56.78","dns":"192.168.1.254","usingStaticIp":false},{"interface":"WAN 2","status":"Not connected"}]'
				uplink_split = uplink_curl.split(',')
				#print("\nuplink_split: ",uplink_split)
				#print()
				#print()
				
				clean_uplink_list = list() # create a new list for cleaned up text
				uplink_section_number = 0 # create a section variable.  This helps with keeping the devices separate.
		
				# clear out uplink variables:
				uplink_interface = "blank"
				uplink_status = "blank"
				investigate_uplink = "no" # if this is set to yes, the device info will be output/sent to the alert.
		
				for uplink_curl_entry in uplink_split:
					#print("uplink_curl_entry Number: ",uplink_section_number)
					uplink_section_number += 1 # update the section counter
					#print("investigate_uplink (within uplink_split loop): ",investigate_uplink)
					#print()
					#print()
					uplink_device_list = list() # create a new list for each entry
					
					#print("uplink_curl_entry (pre-cleanup and split):\n",uplink_curl_entry) # for testing
					#print() # blank line for testing
					uplink_curl_entry = uplink_curl_entry.replace("{","") # remove {
					uplink_curl_entry = uplink_curl_entry.replace("}","") # remove }
					uplink_curl_entry = uplink_curl_entry.replace("]","") # remove ]
					uplink_curl_entry = uplink_curl_entry.replace("[","") # remove [
					uplink_curl_entry = uplink_curl_entry.replace("b'","") # remove b'
					uplink_curl_entry = uplink_curl_entry.replace('"',"") # remove "
					uplink_curl_entry = uplink_curl_entry.replace("'","") # remove '
					#print("uplink_curl_entry after no {} and whatnot: ",uplink_curl_entry) # used for testing
					#print() # blank line for testing
					
					uplink_device_list = uplink_curl_entry.split(',')
					#print("uplink_device_list after cleanup and split:\n",uplink_device_list)
					#print("\n\n\n\n")
					
					for uplink_device_entry in uplink_device_list:
						if investigate_uplink == 'yes':
							#print("break out of loop.")
							break # break out of the second loop since wan1 (previous interface) is flagged.
						#print("***uplink_device_entry: ",uplink_device_entry) # test output: "***uplink_device_entry:  interface:WAN 1" or "***uplink_device_entry:  status:Not connected" etc.
						uplink_attribute = uplink_device_entry.split(':')
						#print("uplink_attribute: ",uplink_attribute) # test output: uplink_attribute:  ['interface', 'WAN 1']
						#print("investigate_uplink (within uplink_device_list loop): ",investigate_uplink)
						# start grabbing data (interface names, values, etc)
						if uplink_attribute[0] == 'interface':
							uplink_interface = uplink_attribute[1] # grab the interface name (WAN 1 or WAN 2)
						if uplink_attribute[0] == 'status':
							uplink_status = uplink_attribute[1] # grab the interface status (Active or Failed or Not connected) - not sure what the options are.
							#print("uplink_status: ",uplink_status)
							if uplink_status != 'Active': 
								if uplink_status != 'Ready':
									#print("uplink_status (when conditions are met. uplink_status != Active or Ready.): ",uplink_status)
									#print("investigate_uplink (uplink_status != Active or Ready.  Before 'investigate_uplink' is set to yes): ",investigate_uplink)
									investigate_uplink = 'yes'
									time = datetime.now() # set the time.  This will be used later to compare to a previous entry.
									#print("investigate_uplink (uplink_status != Active or Ready.  After 'investigate_uplink' is set to yes): ",investigate_uplink)
									break
							elif uplink_status is 'Ready' or uplink_status is 'Active':
								#print("uplink_status is 'Ready' or uplink_status is 'Active': ",uplink_status)
								continue
						
					
					#print("investigate_uplink (End of uplink_split loop): ",investigate_uplink)
					#print("\n\n")
				
				if investigate_uplink == 'no':	# check the file to see if the entry is in the list.  if it is, remove the line since the uplink no longer has issues.
					print("Checking for old entry and removing it if it exists:")
					print("sn_name (Device name): ",sn_name)
					sheet_row_count = get_row_count()
					print("sheet_row_count: ",sheet_row_count)
					if sheet_row_count == 1:
						#print("Sheet only has the headers, no entries to check.")
						continue
					elif sheet_row_count > 1: # if there are more lines than just the headers...
						#print("sheet_row_count has more than 1 row.  It has: ",sheet_row_count)
						for excel_line_num in range(2,sheet_row_count+1): # Skip the first line and go through the first column and search for the device name.  If it exists, delete the line.
							#print("excel_line_num: ",excel_line_num)
							#print("get_row_count()+1: ",get_row_count()+1)
							#print("ws.cell(row=excel_line_num, column=1).value = ",ws.cell(row=excel_line_num, column=1).value)
							#print("sn_name = ",sn_name)
							if ws.cell(row=excel_line_num, column=1).value == sn_name: # if the line exists, delete it.
								ws.delete_rows(excel_line_num)
								#print("Deleting row ",excel_line_num,"...")
								break
							else: # if the device isn't in the sheet, continue
								print("Entry doesn't exist in this row (",excel_line_num,").  Continuing...")
								continue
					
					try: # try to save the sheet
						save_xlsx() # save the xlsx file.
					except PermissionError: # output an error if you get a permission error.  Usually caused by the file being open.  Consider backing up the original for viewing while the script is running.  Won't be used for live updates though.
						print(uplink_excel_list+" is currently open.  Close the file and rerun the script.")
				
				if investigate_uplink == 'yes': # Print the device info with uplinks that are not 'Active'
					print("************************************") 
					print("Device with a uplink to investigate:")
					print("sn_name (Device name): ",sn_name)
					#print("sn_serial (Serial number): ",sn_serial)
					#print("sn_model (Model): ",sn_model)
					#print("uplink_interface (Interface name): ",uplink_interface)
					#print("uplink_status (Int status): ",uplink_status)
					#print("Date / Time: ",time) 
					print("************************************\n")
					investigate_uplink = "no" # reset the variable
			
#################################################################################
#################################################################################
#		save the results to an excel file to track timestamps.				    #
#################################################################################
					
					print("starting to read/write excel file values...\n")
					#print("get_row_count(): ",get_row_count())
					#print("range(1,get_row_count()): ",range(1,get_row_count()))
					sheet_row_count = get_row_count()
					#print("sheet_row_count: ",sheet_row_count)
					if sheet_row_count == 1:
						sheet_row_count += 1
					#print("sheet_row_count (after 1 row check): ",sheet_row_count)
					sheet_rows = range(1,sheet_row_count+1)
					#print("range(1,get_row_count()+1): ",range(1,get_row_count()+1))
					#print("sheet_rows: ",sheet_rows)
					for excel_line_num in sheet_rows:
						#print("\nexcel_line_num: ",excel_line_num)
						#print("get_row_count(): ",get_row_count())
						if excel_line_num == 1: # skip the first row since it's a header row.
							#print("Skipping first row (Headers)\n")
							continue
						#print("ws.cell(row=excel_line_num, column=1).value: ",ws.cell(row=excel_line_num, column=1).value)
						#print("sn_name: ",sn_name)
						
						#if ws.cell(row=excel_line_num, column=1).value != sn_name: # if sn_name isn't found in column A, go to the next line.  
						#	print("\nsn_name not found, in this row.  Checking next row")
						if ws.cell(row=excel_line_num, column=1).value == sn_name: # if the line exists, check times in line 6 and 7
							#if the cell in column 7 is empty; write a new timestamp.
							print("\nEntry "+sn_name+" exists, overwriting column 7.")
							ws.cell(column=7, row=excel_line_num, value=time) # write the time
							original_time = ws.cell(column=6, row=excel_line_num).value # set the original time from column 6 (time when the entry was first wrote to the sheet)
							#print("original_time: ",original_time)
							recent_time = ws.cell(column=7, row=excel_line_num).value # set the newest timestamp.
							#print("recent_time: ",recent_time)
							time_diff = recent_time - original_time # check for a difference
							#print("time_diff: ",time_diff)
							converted_time_diff = time_diff.total_seconds() / 60 # convert the difference to minutes.
							#print("converted_time_diff (minutes): ",converted_time_diff)
							if converted_time_diff >= 45: # if the time difference is over 45 minutes...
								#print("The amount of time that this uplink has been offline is too d4mn high... Adding to alert.")
								ws.cell(column=8, row=excel_line_num).value = 1 # update the column to be added to the alert 
								
							#else:
								#print("Time difference condition has not been met.")
							break # exit the loop and move on to the next device--if any
						elif ws.cell(row=excel_line_num, column=1).value != sn_name and excel_line_num != get_row_count():
							#print("No match in the excel row check. This is not the last row.  Continue to next row")
							continue
						else:
							print("no match in the excel row check. (ws.cell(row=excel_line_num, column=1).value != sn_name)")
							print("creating new line...")
							#if excel_line_num > 2:
							#	excel_line_num += 1 # add a new row
							if ws.cell(column=1, row=excel_line_num).value != None:
								excel_line_num += 1 # add a new row
							print("Adding sn_name: "+sn_name+" in cell: ws A",excel_line_num)
							ws.cell(column=1, row=excel_line_num, value=sn_name) # write the name entry
							ws.cell(column=2, row=excel_line_num, value=sn_serial) # write the SN entry
							ws.cell(column=3, row=excel_line_num, value=sn_model) # write the model
							ws.cell(column=4, row=excel_line_num, value=uplink_interface) # write the uplink interface
							ws.cell(column=5, row=excel_line_num, value=uplink_status) # write the uplink status
							#print("writing time: ws.cell(column=6, row=excel_line_num, value=",time,")")
							ws.cell(column=6, row=excel_line_num, value=time) # write the time
							break # exit the loop and move on to the next device--if any
							

							
					try: # try to save the sheet
						save_xlsx() # save the xlsx file.
					except PermissionError: # output an error if you get a permission error.  Usually caused by the file being open.  Consider backing up the original for viewing while the script is running.  Won't be used for live updates though.
						print(uplink_excel_list+" is currently open.  Close the file and rerun the script.")		
					
if os.path.isfile(alert_excel_list) == True: # check if the REPORT excel file already exists.  If it does, delete it.  This is currently more simple than going through the list and deleting old lines.
	print(alert_excel_list+" does exist.  Deleting it now...")
	os.remove(alert_excel_list)
	print("Delete complete.")
if os.path.isfile(alert_excel_list) == False: # check if the REPORT excel file already exists.  If it doesn't, create the file and write the headers.
	print()
	print(alert_excel_list+" does not exist.  Creating it now")
	wb_report = Workbook() # create workbook
	ws_report = wb_report.active # grab the sheet
	ws_report.cell(column=1, row=1, value="Device name") # write column headers
	ws_report.cell(column=2, row=1, value="Serial number")
	ws_report.cell(column=3, row=1, value="Model")
	ws_report.cell(column=4, row=1, value="Interface")
	ws_report.cell(column=5, row=1, value="Int Status")
	ws_report.cell(column=6, row=1, value="Date Time1")
	ws_report.cell(column=7, row=1, value="Date Time2") # this second column is to compare with the first.  if the difference is met, then the device info is sent in an alert.
	ws_report.cell(column=8, row=1, value="Alert?")
	print("Created file and wrote headers...")
	save_xlsx()

wb.save(alert_excel_list) # Save the alert workbook

# go through the list and anything w/ a 1 in column '8' will be added to a sheet/report that will be emailed to the team.
print("Reading the final list, checking for entries to send an alert on.")
final_sheet_row = get_row_count() # get the row count of the devices that have down uplinks.  We'll go through that and search for anything that's marked for an alert.
for i in range(2,final_sheet_row+1):
	print("Checking uplink links row: ",i)
	#print("ws.cell(column=1, row=i).value: ",ws.cell(column=1, row=i).value)
	#print("ws.cell(column=8, row=i).value: ",ws.cell(column=8, row=i).value)
	if ws.cell(column=8, row=i).value is 1:
		report_row_count = ws_report.max_row # get the max row count of the report file
		#print("\nreport_row_count in the alert sheet: ",report_row_count)
		#print("range(2,report_row_count+1): ",range(2,report_row_count+2))
		if report_row_count is 1:
			report_row_count+=2
			#print("report_row_count+=2")
			#print("\nreport_row_count in the alert sheet: ",report_row_count)
		else:
			report_row_count+=1
			#print("report_row_count+=2")
			#print("\nreport_row_count in the alert sheet: ",report_row_count)
		for y in range(2,report_row_count): # go through the new report and skip the header row.
			print("\ny = ",y)
			print("report_row_count: ",report_row_count)
			if ws_report.cell(column=1, row=y).value is ws.cell(column=1, row=i).value:
				print("Device ",ws.cell(column=1, row=i).value," is already in the report list on row y: ",y)
				print("Overwriting column 7, row ",y," since the entry already is in the list.")
				ws_report.cell(column=7, row=y).value = ws.cell(column=7, row=i).value
				wb.save(alert_excel_list) # Save the alert workbook
				break
			else:
				print("Device entry not found on row: ",y)
			if y is report_row_count+1: # if it's the last row and the entry isn't found, add it to the end of the report.
				print("Device entry not found on row: ",y,".  Writing to new row.")
				new_report_row = y
				print("new_report_row: ",new_report_row,".  Writing values...\n\n")
				ws_report.cell(column=1, row=new_report_row).value = ws.cell(column=1, row=i).value
				ws_report.cell(column=2, row=new_report_row).value = ws.cell(column=2, row=i).value
				ws_report.cell(column=3, row=new_report_row).value = ws.cell(column=3, row=i).value
				ws_report.cell(column=4, row=new_report_row).value = ws.cell(column=4, row=i).value
				ws_report.cell(column=5, row=new_report_row).value = ws.cell(column=5, row=i).value
				ws_report.cell(column=6, row=new_report_row).value = ws.cell(column=6, row=i).value
				ws_report.cell(column=7, row=new_report_row).value = ws.cell(column=7, row=i).value
				wb.save(alert_excel_list) # Save the alert workbook
				break
				
		wb.save(alert_excel_list) # Save the alert workbook
		
# send an email, with the alert sheet, to the team.
subprocess.call("EmailUplinkReport.py", shell=True)